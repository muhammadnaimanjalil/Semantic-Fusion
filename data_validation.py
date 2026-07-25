"""Dataset validation for uploaded machine-learning tables."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Literal

import numpy as np
import pandas as pd
from openpyxl import load_workbook

Severity = Literal["error", "warning"]

EXCEL_ERROR_TOKENS = {
    "#CALC!",
    "#DIV/0!",
    "#FIELD!",
    "#GETTING_DATA",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#SPILL!",
    "#VALUE!",
}
MAX_EXCEL_SCAN_CELLS = 2_000_000
MAX_EXAMPLES = 8


@dataclass(frozen=True)
class DataQualityIssue:
    severity: Severity
    code: str
    message: str
    affected_count: int
    columns: tuple[str, ...] = ()
    examples: tuple[str, ...] = ()


@dataclass(frozen=True)
class DataQualityReport:
    issues: tuple[DataQualityIssue, ...]

    @property
    def errors(self) -> tuple[DataQualityIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == "error")

    @property
    def warnings(self) -> tuple[DataQualityIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == "warning")

    @property
    def can_proceed(self) -> bool:
        return not self.errors


def combine_reports(*reports: DataQualityReport) -> DataQualityReport:
    return DataQualityReport(
        issues=tuple(
            issue
            for report in reports
            for issue in report.issues
        )
    )


def validate_dataframe(frame: pd.DataFrame) -> DataQualityReport:
    issues: list[DataQualityIssue] = []
    if frame.empty:
        issues.append(
            DataQualityIssue(
                severity="error",
                code="empty_dataset",
                message="The uploaded dataset contains no data rows.",
                affected_count=0,
            )
        )
        return DataQualityReport(tuple(issues))
    if len(frame.columns) == 0:
        issues.append(
            DataQualityIssue(
                severity="error",
                code="no_columns",
                message="The uploaded dataset contains no columns.",
                affected_count=0,
            )
        )
        return DataQualityReport(tuple(issues))

    column_names = [str(column) for column in frame.columns]
    duplicate_names = sorted(
        {name for name in column_names if column_names.count(name) > 1}
    )
    if duplicate_names:
        issues.append(
            DataQualityIssue(
                severity="error",
                code="duplicate_columns",
                message="Column names must be unique.",
                affected_count=len(duplicate_names),
                columns=tuple(duplicate_names[:MAX_EXAMPLES]),
            )
        )

    blank_names = [
        name for name in column_names if not name.strip() or name.startswith("Unnamed:")
    ]
    if blank_names:
        issues.append(
            DataQualityIssue(
                severity="error",
                code="blank_columns",
                message="Every column must have a meaningful header.",
                affected_count=len(blank_names),
                columns=tuple(blank_names[:MAX_EXAMPLES]),
            )
        )

    all_missing = [
        str(column) for column in frame.columns if frame[column].isna().all()
    ]
    if all_missing:
        issues.append(
            DataQualityIssue(
                severity="error",
                code="all_missing_columns",
                message="Remove or populate columns that contain no values.",
                affected_count=len(all_missing),
                columns=tuple(all_missing[:MAX_EXAMPLES]),
            )
        )

    numeric = frame.select_dtypes(include=[np.number])
    if not numeric.empty:
        infinite_mask = np.isinf(numeric.to_numpy(dtype=float, copy=False))
        infinite_count = int(infinite_mask.sum())
        if infinite_count:
            infinite_columns = tuple(
                str(column)
                for column in numeric.columns[
                    np.any(infinite_mask, axis=0)
                ][:MAX_EXAMPLES]
            )
            issues.append(
                DataQualityIssue(
                    severity="error",
                    code="infinite_values",
                    message="Replace positive or negative infinity with finite values.",
                    affected_count=infinite_count,
                    columns=infinite_columns,
                )
            )

        zero_counts = (numeric == 0).sum()
        zero_counts = zero_counts[zero_counts > 0]
        if not zero_counts.empty:
            examples = tuple(
                f"{column}: {int(count):,}"
                for column, count in zero_counts.head(MAX_EXAMPLES).items()
            )
            issues.append(
                DataQualityIssue(
                    severity="warning",
                    code="zero_values",
                    message=(
                        "Numeric zero values were found. Zero can be valid, but "
                        "confirm it does not represent missing or uncoded data."
                    ),
                    affected_count=int(zero_counts.sum()),
                    columns=tuple(map(str, zero_counts.index[:MAX_EXAMPLES])),
                    examples=examples,
                )
            )

    object_columns = frame.select_dtypes(include=["object", "string"]).columns
    error_locations: list[str] = []
    error_count = 0
    for column in object_columns:
        normalized = frame[column].dropna().astype(str).str.strip().str.upper()
        mask = normalized.isin(EXCEL_ERROR_TOKENS)
        error_count += int(mask.sum())
        if mask.any() and len(error_locations) < MAX_EXAMPLES:
            error_locations.extend(
                f"{column} (row {index + 2})"
                for index in normalized.index[mask][
                    : MAX_EXAMPLES - len(error_locations)
                ]
            )
    if error_count:
        issues.append(
            DataQualityIssue(
                severity="error",
                code="spreadsheet_errors",
                message="Spreadsheet error values must be corrected before analysis.",
                affected_count=error_count,
                examples=tuple(error_locations),
            )
        )

    missing_counts = frame.isna().sum()
    missing_counts = missing_counts[missing_counts > 0]
    if not missing_counts.empty:
        examples = tuple(
            f"{column}: {int(count):,}"
            for column, count in missing_counts.head(MAX_EXAMPLES).items()
        )
        issues.append(
            DataQualityIssue(
                severity="warning",
                code="missing_values",
                message=(
                    "Missing cells were found. Predictors will be imputed and "
                    "rows with missing targets will be excluded."
                ),
                affected_count=int(missing_counts.sum()),
                columns=tuple(map(str, missing_counts.index[:MAX_EXAMPLES])),
                examples=examples,
            )
        )

    duplicate_count = int(frame.duplicated().sum())
    if duplicate_count:
        issues.append(
            DataQualityIssue(
                severity="warning",
                code="duplicate_rows",
                message="Exact duplicate rows were found; verify they are intentional.",
                affected_count=duplicate_count,
            )
        )

    constant_columns = [
        str(column)
        for column in frame.columns
        if not frame[column].isna().all()
        and frame[column].nunique(dropna=True) <= 1
    ]
    if constant_columns:
        issues.append(
            DataQualityIssue(
                severity="warning",
                code="constant_columns",
                message=(
                    "Constant columns add no predictive information and should "
                    "normally be excluded."
                ),
                affected_count=len(constant_columns),
                columns=tuple(constant_columns[:MAX_EXAMPLES]),
            )
        )
    return DataQualityReport(tuple(issues))


def inspect_excel_sheet(file_bytes: bytes, sheet_name: str) -> DataQualityReport:
    """Detect formulas and stored Excel errors before pandas converts them."""
    issues: list[DataQualityIssue] = []
    workbook = load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=False,
    )
    try:
        worksheet = workbook[sheet_name]
        used_cells = worksheet.max_row * worksheet.max_column
        if used_cells > MAX_EXCEL_SCAN_CELLS:
            issues.append(
                DataQualityIssue(
                    severity="warning",
                    code="limited_formula_scan",
                    message=(
                        "The worksheet is too large for a complete formula scan. "
                        "Only tabular value checks will be applied."
                    ),
                    affected_count=used_cells,
                )
            )
            return DataQualityReport(tuple(issues))

        formula_examples: list[str] = []
        error_examples: list[str] = []
        formula_count = 0
        error_count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formula_count += 1
                    if len(formula_examples) < MAX_EXAMPLES:
                        formula_examples.append(cell.coordinate)
                elif cell.data_type == "e":
                    error_count += 1
                    if len(error_examples) < MAX_EXAMPLES:
                        error_examples.append(
                            f"{cell.coordinate}: {cell.value}"
                        )
        if formula_count:
            issues.append(
                DataQualityIssue(
                    severity="error",
                    code="excel_formulas",
                    message=(
                        "Formula cells were found. Upload a value-only dataset so "
                        "model inputs are explicit and reproducible."
                    ),
                    affected_count=formula_count,
                    examples=tuple(formula_examples),
                )
            )
        if error_count:
            issues.append(
                DataQualityIssue(
                    severity="error",
                    code="excel_errors",
                    message="Excel error cells must be corrected before analysis.",
                    affected_count=error_count,
                    examples=tuple(error_examples),
                )
            )
    finally:
        workbook.close()
    return DataQualityReport(tuple(issues))
